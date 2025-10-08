import time
from datetime import datetime
import os
import openpyxl
from openpyxl import load_workbook
import sys
from playwright.sync_api import Page, TimeoutError

def crear_directorio_si_no_existe(ruta):
    if not os.path.exists(ruta):
        os.makedirs(ruta)

def definirRuta(base_path, complementoRuta=""):
    ruta = os.path.join(base_path, complementoRuta.strip("/\\"))
    return ruta.replace('\\', '/')


def crearArchivoAuditoria(rutaAuditoria):
    fechaAuditoria = obtenerFechaActual()
    Auditoria = rutaAuditoria+f'/Auditoria-{fechaAuditoria}.xlsx'
    filesheet = Auditoria
    wb = openpyxl.Workbook()
    sheet = wb.active
    rowExcel=1 
    rowExcelString=str(rowExcel)
    sheet['A'+rowExcelString] = "INSTANCIA"
    sheet['B'+rowExcelString] = "OBSERVACION"
    sheet['C'+rowExcelString] = "FECHA"
    wb.save(filesheet)
    
    return Auditoria

def diligenciarAuditoria(archivoAuditoria, datos, fila):
    fecha = obtenerFechaActual()
    wb = load_workbook(archivoAuditoria)
    sheet = wb.active
    rowAuditString=str(fila)
    sheet['A'+rowAuditString] = datos[0]
    sheet['B'+rowAuditString] = datos[1]
    sheet['C'+rowAuditString] = fecha
    wb.save(archivoAuditoria)
    
    #tomar pantallazó de error
    #tomar_captura_pantalla(fecha, imagenesAuditoria)
    
    return fila+1

def formatearFecha(fecha):
    print(fecha)  
    fecha = str(fecha)
    fechaSinHora = fecha.replace('00:00:00','')
    fechaSinSlash = fechaSinHora.replace('-','')
    print(fechaSinSlash)
    return fechaSinSlash


def obtenerFechaActual():
    fechaAuditoria = time.strftime("%d-%h-%Y-%H-%M-%p")
    return fechaAuditoria

def eliminarArchivo(rutaArchivo):
    if os.path.exists(rutaArchivo):
        os.remove(rutaArchivo)
    else:
        print(f"El archivo {rutaArchivo} no existe.")

def seleccionar_personalizado(page: Page):
    try:
        # 1. Abrir el ng-select
        page.locator("ng-select").nth(0).click()
        page.wait_for_selector("ng-dropdown-panel", timeout=3000)

        # 2. Intentar seleccionar "Personalizado"
        try:
            page.locator("div.ng-option span", has_text="Personalizado").click(timeout=5000)
            print("Opción 'Personalizado' seleccionada.")
            return True
        except TimeoutError:
            pass  # continúa a buscar "Custom"
        except Exception:
            pass

        # 3. Intentar seleccionar "Custom"
        try:
            page.locator("div.ng-option span", has_text="Custom").click(timeout=5000)
            print("Opción 'Custom' seleccionada.")
            return True
        except Exception as e:
            print(f"No se encontró la opción 'Custom' tampoco: {e}")
            return False

    except Exception as e:
        print(f"Error al interactuar con el dropdown: {e}")
        return False



def adaptar_formato_fecha(instancia: str, fecha_inicio: str, fecha_fin: str):
    """Adapta el formato de fecha según la instancia"""
    # Convertir las fechas a objeto datetime
    fecha_inicio_dt = datetime.strptime(fecha_inicio, "%m/%d/%Y")
    fecha_fin_dt = datetime.strptime(fecha_fin, "%m/%d/%Y")

    # Si la instancia es la que usa formato DD/MM/YYYY
    if instancia == "Instancia 3":
        formato = "%d/%m/%Y"
    else:
        formato = "%m/%d/%Y"

    # Retornar las fechas en el formato requerido
    return (
        fecha_inicio_dt.strftime(formato),
        fecha_fin_dt.strftime(formato),
    )

def limpiar_vacios(df, columnas=None):
    """
    Limpia valores NaN, None, null o 'nan' en un DataFrame.
    Si columnas es None, se aplicará a todo el DataFrame.
    """
    if columnas is None:
        columnas = df.columns

    # Reemplazar NaN/None con vacío antes de convertir a str
    df[columnas] = df[columnas].fillna("").astype(str)

    # Reemplazar cadenas 'nan', 'None', 'null' por vacío
    df[columnas] = df[columnas].replace(
        to_replace=["nan", "None", "null", "NaT"],
        value="",
        regex=False
    )

    return df
